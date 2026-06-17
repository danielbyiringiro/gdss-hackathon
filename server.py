"""
Product Attribute Extraction Server  (image -> Item Master Database)

Two interchangeable vision backends:
  * "claude"  -> Anthropic Claude vision API   (default, most accurate)
  * "ollama"  -> local Ollama + LLaVA          (offline fallback)

Choose with the BACKEND env var, or per-request with the `backend` form field.
Serves a web UI at  http://localhost:8000/  and a REST API at /extract.
"""

import asyncio
import base64
import json
import os
import re
import threading

import httpx
import uvicorn
from fastapi import FastAPI, File, Form, HTTPException, UploadFile

# Lightweight .env loader (no dependency): populate os.environ from ./.env
_envf = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_envf):
    for _line in open(_envf, encoding="utf-8"):
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            _v = _v.strip()
            # strip an inline comment ( value  # comment ) unless quoted
            if _v[:1] not in ('"', "'"):
                _v = re.split(r"\s+#", _v, maxsplit=1)[0]
            os.environ.setdefault(_k.strip(), _v.strip().strip('"').strip("'"))
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from openpyxl import Workbook, load_workbook

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------
def _auto_backend():
    if os.environ.get("OPENAI_API_KEY"):
        return "openai"
    if os.environ.get("ANTHROPIC_API_KEY"):
        return "claude"
    return "ollama"

DEFAULT_BACKEND = os.environ.get("BACKEND", _auto_backend())

# Ollama (local LLaVA)
OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "llava")

# Claude (Anthropic API)
ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")

# OpenAI (GPT-4o vision)
OPENAI_URL = "https://api.openai.com/v1/chat/completions"
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-4o")

FIELDS = [
    "ITEM_NAME", "BARCODE", "MANUFACTURER", "BRAND", "WEIGHT",
    "PACKAGING_TYPE", "COUNTRY", "VARIANT", "TYPE", "FRAGRANCE_FLAVOR",
    "PROMOTION", "ADDONS", "TAGLINE",
]

FIELD_DESCRIPTIONS = """
- ITEM_NAME: Full catalog-style product name combining brand, variant/flavor, weight, packaging, type, manufacturer and country (e.g. "BLUE BAND SALTED MARGARINE 250G TUB UPFIELD GHANA")
- BARCODE: The numeric digits printed below the barcode lines; 8, 12, 13 or 14 digits only, no spaces. "" if not clearly readable.
- MANUFACTURER: Company that manufactures or is responsible for the product
- BRAND: Brand name as shown on the package
- WEIGHT: Net weight or volume with unit, exactly as on pack (e.g. "250G", "330ML", "2200G")
- PACKAGING_TYPE: Physical container type, uppercase (e.g. "TUB", "SACHET", "BOTTLE", "CAN", "GLASS JAR", "BOX", "TIN", "CARTON")
- COUNTRY: Country of manufacture or packing
- VARIANT: Product variant (e.g. "ORIGINAL", "3 IN 1", "PREMIUM"); "" if not applicable
- TYPE: Short product category (e.g. "MARGARINE", "MAYONNAISE", "TOMATO MIX", "SOFT DRINK", "SEASONING POWDER")
- FRAGRANCE_FLAVOR: Flavor or fragrance (e.g. "ROSE", "STRAWBERRY", "CITRUS"); "" if not applicable
- PROMOTION: On-pack promotion text (e.g. "25+7 FREE", "BUY NOW GHS33"); "" if none
- ADDONS: Extra pack contents / features (e.g. "STRAW INCLUDED", "100 TEABAGS"); "" if none
- TAGLINE: Short promotional tagline; "" if none
"""

EXTRACTION_PROMPT = f"""You are a product data extraction specialist building an Item Master Database.
Read ALL text on this product packaging image carefully, including small print, and extract:
{FIELD_DESCRIPTIONS}

Rules:
- Only extract what is clearly visible. Do NOT guess or hallucinate.
- BARCODE: return the digits only if you can read them confidently (8/12/13/14 digits), else "".
- WEIGHT: keep the on-pack format, uppercase, no space ("250G" not "250 g").
- Use uppercase for short categorical fields.
- Return "" for any field not visible/applicable.

Return ONLY a valid JSON object with exactly these keys: {json.dumps(FIELDS)}
No explanation, no markdown — just the raw JSON object."""

OUTPUT_XLSX = os.environ.get("OUTPUT_XLSX", "output_results.xlsx")
_xlsx_lock = threading.Lock()

app = FastAPI(title="Product Attribute Extractor (image -> IMDB)")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


# --------------------------------------------------------------------------
# Parsing / cleaning helpers
# --------------------------------------------------------------------------
def sanitize_barcode(val) -> str:
    digits = re.sub(r"\D", "", str(val or ""))
    if not digits or len(digits) not in (8, 12, 13, 14):
        return ""
    if re.fullmatch(r"0+", digits) or len(set(digits)) == 1:
        return ""
    if re.fullmatch(r"(0123456789|1234567890)+\d{0,3}", digits):
        return ""
    return digits


def clean_json(text: str) -> dict:
    text = re.sub(r"```json|```", "", (text or "").strip()).strip()
    text = re.sub(r"\d{20,}", lambda m: m.group(0)[:20], text)
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if match:
        text = match.group()
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        repaired = text
        if repaired.count('"') % 2 == 1:
            repaired += '"'
        open_braces = repaired.count("{") - repaired.count("}")
        if open_braces > 0:
            repaired += "}" * open_braces
        parsed = json.loads(repaired)

    result = {}
    for field in FIELDS:
        val = (parsed.get(field) or parsed.get(field.lower())
               or parsed.get(field.replace("_", " ")) or "")
        val = str(val).strip() if val else ""
        if field == "BARCODE":
            val = sanitize_barcode(val)
        result[field] = val
    return result


def append_to_xlsx(product: dict, path: str = None) -> int:
    path = path or OUTPUT_XLSX
    with _xlsx_lock:
        if os.path.exists(path):
            wb = load_workbook(path); ws = wb.active
        else:
            wb = Workbook(); ws = wb.active; ws.title = "Sheet1"; ws.append(FIELDS)
        row = []
        for field in FIELDS:
            val = product.get(field, "")
            if field == "BARCODE" and val:
                try:
                    row.append(int(val)); continue
                except (ValueError, TypeError):
                    pass
            row.append(val if val else None)
        ws.append(row); wb.save(path)
        return ws.max_row


# --------------------------------------------------------------------------
# Backend: Claude vision (Anthropic API)
# --------------------------------------------------------------------------
async def extract_claude(image_bytes: bytes, mime_type: str) -> dict:
    if not ANTHROPIC_API_KEY:
        raise HTTPException(503, "ANTHROPIC_API_KEY not set. Export it or use BACKEND=ollama.")
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    payload = {
        "model": CLAUDE_MODEL,
        "max_tokens": 1024,
        "temperature": 0,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64",
                 "media_type": mime_type or "image/jpeg", "data": b64}},
                {"type": "text", "text": EXTRACTION_PROMPT},
            ],
        }],
    }
    headers = {"x-api-key": ANTHROPIC_API_KEY,
               "anthropic-version": "2023-06-01",
               "content-type": "application/json"}
    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post(ANTHROPIC_URL, headers=headers, json=payload)
        resp.raise_for_status()
        data = resp.json()
    raw = "".join(b.get("text", "") for b in data.get("content", []))
    try:
        return clean_json(raw)
    except Exception as e:
        print(f"Claude parse error: {e}\nRaw: {raw[:300]}")
        return {f: "" for f in FIELDS}


# --------------------------------------------------------------------------
# Backend: OpenAI GPT-4o vision
# --------------------------------------------------------------------------
OPENAI_DETAIL = os.environ.get("OPENAI_DETAIL", "high")  # high | low | auto


def _retry_wait(resp, attempt) -> float:
    """Seconds to wait before retrying a 429, from the API hint or backoff."""
    ra = resp.headers.get("retry-after")
    if ra:
        try:
            return float(ra) + 0.3
        except ValueError:
            pass
    m = re.search(r"try again in ([\d.]+)(ms|s)", resp.text)
    if m:
        secs = float(m.group(1)) / (1000 if m.group(2) == "ms" else 1)
        # Floor at 4s: per-minute token windows rarely clear in <1s, and
        # retrying too soon just wastes an attempt on another 429.
        return max(secs + 0.3, 4.0)
    return min(2 ** attempt, 45)

async def extract_openai(image_bytes: bytes, mime_type: str) -> dict:
    if not OPENAI_API_KEY:
        raise HTTPException(503, "OPENAI_API_KEY not set. Export it or pick another backend.")
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    data_uri = f"data:{mime_type or 'image/jpeg'};base64,{b64}"
    payload = {
        "model": OPENAI_MODEL,
        "max_tokens": 1024,
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "messages": [{
            "role": "user",
            "content": [
                {"type": "text", "text": EXTRACTION_PROMPT},
                {"type": "image_url", "image_url": {"url": data_uri, "detail": OPENAI_DETAIL}},
            ],
        }],
    }
    headers = {"Authorization": f"Bearer {OPENAI_API_KEY}", "content-type": "application/json"}
    async with httpx.AsyncClient(timeout=120.0) as client:
        for attempt in range(14):
            resp = await client.post(OPENAI_URL, headers=headers, json=payload)
            if resp.status_code == 400 and "max_tokens" in resp.text:
                # Newer models want max_completion_tokens instead of max_tokens.
                payload["max_completion_tokens"] = payload.pop("max_tokens")
                continue
            if resp.status_code == 429:
                wait = _retry_wait(resp, attempt)
                print(f"  rate-limited; waiting {wait:.1f}s (attempt {attempt+1}/14)…")
                await asyncio.sleep(wait)
                continue
            break
        if resp.status_code >= 400:
            raise HTTPException(resp.status_code, f"OpenAI API error: {resp.text[:500]}")
        data = resp.json()
    raw = data["choices"][0]["message"]["content"]
    try:
        return clean_json(raw)
    except Exception as e:
        print(f"OpenAI parse error: {e}\nRaw: {raw[:300]}")
        return {f: "" for f in FIELDS}


# --------------------------------------------------------------------------
# Backend: Ollama / LLaVA (local)
# --------------------------------------------------------------------------
async def extract_ollama(image_bytes: bytes, mime_type: str) -> dict:
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    payload = {"model": OLLAMA_MODEL, "prompt": EXTRACTION_PROMPT, "images": [b64],
               "stream": False, "options": {"temperature": 0.1, "num_predict": 512}}
    async with httpx.AsyncClient(timeout=180.0) as client:
        resp = await client.post(OLLAMA_URL, json=payload)
        resp.raise_for_status()
        data = resp.json()
    try:
        return clean_json(data.get("response", ""))
    except Exception as e:
        print(f"Ollama parse error: {e}")
        return {f: "" for f in FIELDS}


_BACKENDS = {"claude": extract_claude, "openai": extract_openai, "ollama": extract_ollama}

async def extract_one(image_bytes, mime, backend):
    fn = _BACKENDS.get(backend, extract_ollama)
    return await fn(image_bytes, mime)


# --------------------------------------------------------------------------
# Merge across images of the same product
# --------------------------------------------------------------------------
def merge_extractions(extractions: list[dict]) -> dict:
    if len(extractions) == 1:
        return extractions[0]
    merged = {f: "" for f in FIELDS}
    for field in FIELDS:
        candidates = [e[field] for e in extractions if e.get(field, "").strip()]
        if not candidates:
            continue
        if field == "BARCODE":
            counts = {}
            for c in candidates:
                counts[c] = counts.get(c, 0) + 1
            merged[field] = max(counts.items(), key=lambda kv: (kv[1], len(kv[0])))[0]
        elif len(set(candidates)) == 1:
            merged[field] = candidates[0]
        else:
            merged[field] = max(candidates, key=len)
    return merged


# --------------------------------------------------------------------------
# API
# --------------------------------------------------------------------------
@app.post("/extract")
async def extract_product(
    images: list[UploadFile] = File(...),
    backend: str = Form(None),
):
    backend = (backend or DEFAULT_BACKEND).lower()
    if not images:
        raise HTTPException(400, "At least one image is required.")
    if len(images) > 12:
        raise HTTPException(400, "Maximum 12 images per request.")

    if backend == "ollama":
        try:
            async with httpx.AsyncClient(timeout=5.0) as c:
                (await c.get("http://localhost:11434/api/tags")).raise_for_status()
        except Exception:
            raise HTTPException(503, "Ollama not running. Start it with: ollama serve")

    extractions, per_image = [], []
    for img in images:
        content = await img.read()
        result = await extract_one(content, img.content_type or "image/jpeg", backend)
        extractions.append(result)
        per_image.append({"filename": img.filename, "extraction": result})

    merged = merge_extractions(extractions)
    row_number = append_to_xlsx(merged)
    model = {"claude": CLAUDE_MODEL, "openai": OPENAI_MODEL}.get(backend, OLLAMA_MODEL)
    return JSONResponse({
        "backend": backend, "model": model,
        "product": merged, "images_processed": len(images), "per_image": per_image,
        "xlsx_row": row_number, "xlsx_path": os.path.abspath(OUTPUT_XLSX),
    })


@app.post("/reset")
async def reset_xlsx():
    with _xlsx_lock:
        if os.path.exists(OUTPUT_XLSX):
            os.remove(OUTPUT_XLSX)
    return {"status": "ok", "message": f"{OUTPUT_XLSX} removed."}


@app.get("/download")
async def download_xlsx():
    if not os.path.exists(OUTPUT_XLSX):
        raise HTTPException(404, "No results yet.")
    return FileResponse(OUTPUT_XLSX, filename="output_results.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/health")
async def health():
    out = {"status": "ok", "default_backend": DEFAULT_BACKEND,
           "openai_key_set": bool(OPENAI_API_KEY), "openai_model": OPENAI_MODEL,
           "claude_key_set": bool(ANTHROPIC_API_KEY), "claude_model": CLAUDE_MODEL}
    try:
        async with httpx.AsyncClient(timeout=4.0) as c:
            r = await c.get("http://localhost:11434/api/tags")
            out["ollama"] = "running"
            out["ollama_models"] = [m["name"] for m in r.json().get("models", [])]
    except Exception:
        out["ollama"] = "not reachable"
    return out


@app.get("/", response_class=HTMLResponse)
async def index():
    here = os.path.dirname(os.path.abspath(__file__))
    with open(os.path.join(here, "index.html"), encoding="utf-8") as f:
        return HTMLResponse(f.read())


if __name__ == "__main__":
    uvicorn.run("server:app", host="0.0.0.0", port=8000, reload=True)
