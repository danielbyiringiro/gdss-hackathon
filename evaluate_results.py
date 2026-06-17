#!/usr/bin/env python3
"""
evaluate_results.py

Compare two product-extraction xlsx files (e.g. a "ground truth" /
reference file and a "predicted" / generated file from server.py) and
report how well they match, using several complementary evaluation
strategies:

  1. Exact match accuracy        - per-field, per-row exact string match
  2. Normalized match accuracy   - case/whitespace/punctuation-insensitive match
  3. Fuzzy similarity (ratio)    - Levenshtein-style similarity score per field
  4. Numeric field tolerance     - BARCODE compared as numbers (ignores
                                    formatting differences)
  5. Empty/missing-value accuracy- did both files agree a field is empty?
  6. Token-set overlap (Jaccard) - useful for long free-text fields like
                                    ITEM_NAME and TAGLINE, where word order
                                    or extra words shouldn't fully penalize
  7. Best-match row alignment    - rows are NOT compared by position. Each
                                    row in file A is matched to the row in
                                    file B with the most similar ITEM_NAME /
                                    BRAND / TYPE / MANUFACTURER / WEIGHT
                                    (greedy bipartite matching), so a missing
                                    or reordered product doesn't cascade into
                                    every later row looking "wrong". Rows
                                    with no good match above --threshold are
                                    reported separately as unmatched.
  8. Per-column and overall summary report

Usage:
  python evaluate_results.py reference.xlsx predicted.xlsx [--out report.xlsx] [--threshold 0.3]
"""

import argparse
import difflib
import re
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("This script requires openpyxl: pip install openpyxl --break-system-packages")
    sys.exit(1)

try:
    from rapidfuzz import fuzz as _rf_fuzz

    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------


def load_sheet(path):
    """Load the active sheet of an xlsx file as (headers, rows)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def align_columns(headers_a, headers_b):
    """
    Build a list of (col_name, idx_a, idx_b) for columns present in both
    files. Column names are matched case-insensitively and with whitespace
    collapsed, so "PACKAGING  TYPE" matches "PACKAGING_TYPE" or
    "Packaging Type".
    """

    def norm(h):
        return re.sub(r"[\s_]+", "_", h.strip().upper())

    norm_a = {norm(h): i for i, h in enumerate(headers_a)}
    norm_b = {norm(h): i for i, h in enumerate(headers_b)}

    common = []
    for key, idx_a in norm_a.items():
        if key in norm_b:
            common.append((headers_a[idx_a], idx_a, norm_b[key]))

    only_a = [headers_a[i] for k, i in norm_a.items() if k not in norm_b]
    only_b = [headers_b[i] for k, i in norm_b.items() if k not in norm_a]

    return common, only_a, only_b


# ---------------------------------------------------------------------------
# Value normalization helpers
# ---------------------------------------------------------------------------


def to_str(val):
    """Convert a cell value to a stripped string, treating None as ''."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def normalize(val):
    """Aggressive normalization: uppercase, collapse whitespace, strip
    common punctuation that doesn't change meaning."""
    s = to_str(val).upper()
    s = re.sub(r"[.,;:!]+", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def is_empty(val):
    s = to_str(val)
    return s == "" or s.lower() == "none"


def to_number(val):
    """Try to parse a value as a number (for BARCODE / weight comparisons)."""
    s = to_str(val)
    s = re.sub(r"[^\d.]", "", s)
    if not s:
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Similarity metrics
# ---------------------------------------------------------------------------


def similarity_ratio(a, b):
    """Character-level similarity in [0, 1]. Uses rapidfuzz if available,
    otherwise difflib (stdlib)."""
    a, b = to_str(a), to_str(b)
    if a == "" and b == "":
        return 1.0
    if a == "" or b == "":
        return 0.0
    if HAVE_RAPIDFUZZ:
        return _rf_fuzz.ratio(a, b) / 100.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def token_jaccard(a, b):
    """Word-level Jaccard similarity in [0, 1]."""
    ta = set(normalize(a).split())
    tb = set(normalize(b).split())
    if not ta and not tb:
        return 1.0
    if not ta or not tb:
        return 0.0
    inter = ta & tb
    union = ta | tb
    return len(inter) / len(union)


# ---------------------------------------------------------------------------
# Per-field comparison strategies
# ---------------------------------------------------------------------------


def is_numeric_field(col_name):
    return "BARCODE" in col_name.upper()


def is_long_text_field(col_name):
    return col_name.upper() in {"ITEM_NAME", "TAGLINE", "ADDONS", "PROMOTION"}


def compare_cell(col_name, val_a, val_b):
    """
    Return a dict of strategy_name -> score (0.0-1.0, or None if N/A) for a
    single cell comparison.
    """
    result = {}

    a_empty, b_empty = is_empty(val_a), is_empty(val_b)

    # 1. Exact match
    result["exact"] = 1.0 if to_str(val_a) == to_str(val_b) else 0.0

    # 2. Normalized match
    result["normalized"] = 1.0 if normalize(val_a) == normalize(val_b) else 0.0

    # 3. Fuzzy character similarity
    result["fuzzy"] = similarity_ratio(val_a, val_b)

    # 4. Numeric tolerance (only meaningful for numeric-looking fields)
    if is_numeric_field(col_name):
        na, nb = to_number(val_a), to_number(val_b)
        if na is None and nb is None:
            result["numeric"] = 1.0 if (a_empty and b_empty) else None
        elif na is None or nb is None:
            result["numeric"] = 0.0
        else:
            result["numeric"] = 1.0 if na == nb else 0.0
    else:
        result["numeric"] = None

    # 5. Empty-agreement
    result["empty_agreement"] = 1.0 if (a_empty == b_empty) else 0.0

    # 6. Token-set Jaccard (most useful for long free-text fields)
    if is_long_text_field(col_name):
        result["jaccard"] = token_jaccard(val_a, val_b)
    else:
        result["jaccard"] = None

    return result


# ---------------------------------------------------------------------------
# Row alignment
# ---------------------------------------------------------------------------

# Fields used to decide whether two rows describe the "same" product,
# in priority order. BARCODE is deliberately NOT used here, since it's
# often unreliable / hallucinated by the extraction model and would cause
# correct rows to be mismatched.
ALIGNMENT_FIELDS = ["ITEM_NAME", "BRAND", "TYPE", "MANUFACTURER", "WEIGHT"]


def row_similarity(common_cols, row_a, row_b):
    """
    Combined similarity score in [0, 1] between two rows, used to decide
    whether they represent the same product. Weighted average of token
    Jaccard + fuzzy similarity over ALIGNMENT_FIELDS (whichever of those
    columns exist in both files).
    """
    cols_by_name = {c[0]: c for c in common_cols}
    scores = []
    weights = []

    for field in ALIGNMENT_FIELDS:
        # column names may have different casing/spacing; find a match
        match_col = None
        for cname in cols_by_name:
            if re.sub(r"[\s_]+", "_", cname.strip().upper()) == field:
                match_col = cols_by_name[cname]
                break
        if match_col is None:
            continue

        _, idx_a, idx_b = match_col
        val_a = row_a[idx_a] if idx_a < len(row_a) else None
        val_b = row_b[idx_b] if idx_b < len(row_b) else None

        if is_empty(val_a) and is_empty(val_b):
            continue

        jac = token_jaccard(val_a, val_b)
        fuz = similarity_ratio(val_a, val_b)
        combined = (jac + fuz) / 2

        # ITEM_NAME is the strongest signal; weight it more
        weight = 3.0 if field == "ITEM_NAME" else 1.0
        scores.append(combined * weight)
        weights.append(weight)

    if not weights:
        return 0.0
    return sum(scores) / sum(weights)


def align_rows(common_cols, rows_a, rows_b, threshold=0.3):
    """
    Greedily match each row in rows_a to its best-matching row in rows_b
    (by row_similarity), so that comparisons are based on which products
    actually correspond, not on row position.

    Returns a list of (idx_a, idx_b_or_None, score) tuples, one per row in
    rows_a. idx_b is None if no row in rows_b scored above `threshold` (and
    wasn't already claimed by a better match).
    """
    n_a, n_b = len(rows_a), len(rows_b)

    # Compute all pairwise similarities
    sims = [
        [row_similarity(common_cols, rows_a[i], rows_b[j]) for j in range(n_b)]
        for i in range(n_a)
    ]

    # Build a sorted list of all (score, i, j) candidates, then greedily
    # assign best matches first (classic greedy bipartite matching).
    candidates = []
    for i in range(n_a):
        for j in range(n_b):
            candidates.append((sims[i][j], i, j))
    candidates.sort(key=lambda x: x[0], reverse=True)

    assigned_a = set()
    assigned_b = set()
    matches = {}

    for score, i, j in candidates:
        if score < threshold:
            break
        if i in assigned_a or j in assigned_b:
            continue
        matches[i] = (j, score)
        assigned_a.add(i)
        assigned_b.add(j)

    result = []
    for i in range(n_a):
        if i in matches:
            j, score = matches[i]
            result.append((i, j, score))
        else:
            result.append((i, None, 0.0))

    unmatched_b = [j for j in range(n_b) if j not in assigned_b]

    return result, unmatched_b


# ---------------------------------------------------------------------------
# Main comparison
# ---------------------------------------------------------------------------


def compare_files(path_a, path_b, threshold=0.3):
    headers_a, rows_a = load_sheet(path_a)
    headers_b, rows_b = load_sheet(path_b)

    common_cols, only_a, only_b = align_columns(headers_a, headers_b)

    alignment, unmatched_b = align_rows(
        common_cols, rows_a, rows_b, threshold=threshold
    )

    n_matched = sum(1 for _, j, _ in alignment if j is not None)

    # per-column aggregate scores
    col_scores = defaultdict(lambda: defaultdict(list))

    # per-row detail (for the report)
    row_details = []
    unmatched_a_rows = []

    for idx_a, idx_b, align_score in alignment:
        row_a = rows_a[idx_a]

        if idx_b is None:
            unmatched_a_rows.append(idx_a + 2)
            continue

        row_b = rows_b[idx_b]
        row_detail = {
            "row_a": idx_a + 2,  # +2 = 1-indexed + header row
            "row_b": idx_b + 2,
            "align_score": round(align_score, 3),
        }

        for col_name, idx_ca, idx_cb in common_cols:
            val_a = row_a[idx_ca] if idx_ca < len(row_a) else None
            val_b = row_b[idx_cb] if idx_cb < len(row_b) else None

            scores = compare_cell(col_name, val_a, val_b)
            for strat, score in scores.items():
                if score is not None:
                    col_scores[col_name][strat].append(score)

            row_detail[col_name] = {
                "a": to_str(val_a),
                "b": to_str(val_b),
                "exact": scores["exact"],
                "fuzzy": round(scores["fuzzy"], 3),
            }

        row_details.append(row_detail)

    return {
        "headers_a": headers_a,
        "headers_b": headers_b,
        "common_cols": [c[0] for c in common_cols],
        "only_a": only_a,
        "only_b": only_b,
        "n_rows_a": len(rows_a),
        "n_rows_b": len(rows_b),
        "n_matched": n_matched,
        "unmatched_a_rows": unmatched_a_rows,
        "unmatched_b_rows": [j + 2 for j in unmatched_b],
        "col_scores": col_scores,
        "row_details": row_details,
    }


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

STRATEGY_LABELS = {
    "exact": "Exact match",
    "normalized": "Normalized match",
    "fuzzy": "Fuzzy similarity (avg)",
    "numeric": "Numeric match",
    "empty_agreement": "Empty/missing agreement",
    "jaccard": "Token overlap (Jaccard)",
}


def avg(values):
    return sum(values) / len(values) if values else None


def print_report(result):
    print("=" * 70)
    print("FILE COMPARISON SUMMARY")
    print("=" * 70)
    print(f"Rows in file A (reference): {result['n_rows_a']}")
    print(f"Rows in file B (predicted): {result['n_rows_b']}")
    print(f"Matched pairs:              {result['n_matched']}")

    if result["only_a"]:
        print(f"\n  Columns only in file A: {result['only_a']}")
    if result["only_b"]:
        print(f"\n  Columns only in file B: {result['only_b']}")

    if result["unmatched_a_rows"]:
        print(
            f"\n  Rows in file A with NO matching row in file B "
            f"(missing / not extracted): {result['unmatched_a_rows']}"
        )

    if result["unmatched_b_rows"]:
        print(
            f"\n  Rows in file B with NO matching row in file A "
            f"(extra / duplicate extractions): {result['unmatched_b_rows']}"
        )

    print("\n" + "-" * 70)
    print("PER-COLUMN SCORES (averaged across compared rows)")
    print("-" * 70)

    header_fmt = "{:<22}" + "".join(f"{{:>16}}" for _ in STRATEGY_LABELS)
    print(header_fmt.format("Column", *STRATEGY_LABELS.values()))

    overall = defaultdict(list)

    for col_name in result["common_cols"]:
        scores = result["col_scores"][col_name]
        row_vals = []
        for strat in STRATEGY_LABELS:
            vals = scores.get(strat, [])
            m = avg(vals)
            if m is not None:
                overall[strat].append(m)
            row_vals.append(f"{m:.1%}" if m is not None else "n/a")
        print(header_fmt.format(col_name, *row_vals))

    print("-" * 70)
    overall_vals = []
    for strat in STRATEGY_LABELS:
        m = avg(overall[strat])
        overall_vals.append(f"{m:.1%}" if m is not None else "n/a")
    print(header_fmt.format("OVERALL", *overall_vals))

    print("\n" + "-" * 70)
    print("WORST-MATCHING ROW PAIRS (by average fuzzy similarity)")
    print("-" * 70)

    row_avgs = []
    for rd in result["row_details"]:
        fuzzy_vals = [rd[c]["fuzzy"] for c in result["common_cols"] if c in rd]
        row_avgs.append((rd["row_a"], rd["row_b"], rd["align_score"], avg(fuzzy_vals)))

    row_avgs.sort(key=lambda x: x[3] if x[3] is not None else 1.0)
    for row_a, row_b, align_score, score in row_avgs[:5]:
        if score is not None:
            print(
                f"  A row {row_a} <-> B row {row_b} (alignment {align_score:.1%}): avg fuzzy similarity = {score:.1%}"
            )
        else:
            print(f"  A row {row_a} <-> B row {row_b}: n/a")


# ---------------------------------------------------------------------------
# Optional detailed xlsx report
# ---------------------------------------------------------------------------


def write_report_xlsx(result, out_path):
    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Column"] + list(STRATEGY_LABELS.values()))

    overall = defaultdict(list)
    for col_name in result["common_cols"]:
        scores = result["col_scores"][col_name]
        row = [col_name]
        for strat in STRATEGY_LABELS:
            vals = scores.get(strat, [])
            m = avg(vals)
            if m is not None:
                overall[strat].append(m)
            row.append(round(m, 4) if m is not None else None)
        ws.append(row)

    overall_row = ["OVERALL"]
    for strat in STRATEGY_LABELS:
        m = avg(overall[strat])
        overall_row.append(round(m, 4) if m is not None else None)
    ws.append(overall_row)

    # --- Per-row detail sheet ---
    ws2 = wb.create_sheet("Row Details")
    header = ["Row A", "Row B", "Alignment score"]
    for col_name in result["common_cols"]:
        header += [
            f"{col_name} (A)",
            f"{col_name} (B)",
            f"{col_name} exact",
            f"{col_name} fuzzy",
        ]
    ws2.append(header)

    for rd in result["row_details"]:
        row = [rd["row_a"], rd["row_b"], rd["align_score"]]
        for col_name in result["common_cols"]:
            cell = rd.get(col_name, {})
            row += [
                cell.get("a", ""),
                cell.get("b", ""),
                cell.get("exact"),
                cell.get("fuzzy"),
            ]
        ws2.append(row)

    # --- Unmatched rows sheet ---
    ws3 = wb.create_sheet("Unmatched Rows")
    ws3.append(["File A rows with no match in B", "File B rows with no match in A"])
    max_len = max(len(result["unmatched_a_rows"]), len(result["unmatched_b_rows"]), 0)
    for i in range(max_len):
        a_val = (
            result["unmatched_a_rows"][i]
            if i < len(result["unmatched_a_rows"])
            else None
        )
        b_val = (
            result["unmatched_b_rows"][i]
            if i < len(result["unmatched_b_rows"])
            else None
        )
        ws3.append([a_val, b_val])

    wb.save(out_path)
    print(f"\nDetailed report written to: {out_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Compare two output_results.xlsx files"
    )
    parser.add_argument("file_a", help="Reference / ground-truth xlsx file")
    parser.add_argument("file_b", help="Predicted / generated xlsx file")
    parser.add_argument("--out", help="Optional path to write a detailed xlsx report")
    parser.add_argument(
        "--threshold",
        type=float,
        default=0.3,
        help="Minimum row-alignment similarity (0-1) to count "
        "as a match (default: 0.3)",
    )
    args = parser.parse_args()

    result = compare_files(args.file_a, args.file_b, threshold=args.threshold)
    print_report(result)

    if args.out:
        write_report_xlsx(result, args.out)


if __name__ == "__main__":
    main()
